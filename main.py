"""
FastAPI backend — pagina pubblica + API chatbot + dashboard admin.
"""
import io
import json
import mimetypes
import re
import shutil
import sqlite3
import threading
import logging
import uuid
from datetime import datetime, timedelta
from pathlib import Path

from fastapi import FastAPI, File, Form, HTTPException, Query, UploadFile
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse, StreamingResponse
from pydantic import BaseModel

from modules.database import (
    init_db, get_complaints, get_complaint_by_id,
    update_complaint, save_complaint,
    get_chatbot_config, save_chatbot_config, get_stats, DB_PATH, UPLOAD_ROOT,
    save_attachment_record, get_pending_attachments, attach_pending_attachment,
    get_complaint_attachments, get_attachment_by_id, delete_complaint,
)
from modules.chatbot import init_state, process_message
from modules.chatbot import build_conversation_history
from modules import llm as llm_module
from modules.constants import COMPLAINT_UPDATABLE_COLUMNS

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(name)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

app = FastAPI(title="San Carlo CarloBot Demo")
init_db()

MAX_UPLOAD_FILES = 10
MAX_IMAGE_BYTES = 8 * 1024 * 1024
ALLOWED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}
MIME_EXTENSION_FALLBACK = {
    "image/jpeg": ".jpg",
    "image/png": ".png",
    "image/gif": ".gif",
    "image/webp": ".webp",
    "image/bmp": ".bmp",
}

# ── Chat session store ────────────────────────────────────────────────────────
_sessions: dict[str, dict] = {}
_lock = threading.Lock()

def get_session(session_id: str) -> dict:
    with _lock:
        if session_id not in _sessions:
            _sessions[session_id] = init_state()
        return dict(_sessions[session_id])

def set_session(session_id: str, state: dict):
    with _lock:
        _sessions[session_id] = state


def _safe_session_dir(session_id: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", session_id or "").strip("._")
    return safe[:80] or "session"


def _safe_original_filename(filename: str | None) -> str:
    name = Path(filename or "immagine").name
    name = re.sub(r"[^\w.\- ()]+", "_", name, flags=re.UNICODE).strip(" ._")
    return name[:120] or "immagine"


def _extension_for_upload(upload: UploadFile) -> str:
    original = _safe_original_filename(upload.filename)
    suffix = Path(original).suffix.lower()
    content_type = (upload.content_type or "").lower()

    if suffix in ALLOWED_IMAGE_EXTENSIONS:
        return suffix
    if content_type in MIME_EXTENSION_FALLBACK:
        return MIME_EXTENSION_FALLBACK[content_type]

    guessed = mimetypes.guess_extension(content_type) if content_type else None
    if guessed and guessed.lower() in ALLOWED_IMAGE_EXTENSIONS:
        return guessed.lower()
    raise HTTPException(status_code=400, detail=f"Formato immagine non supportato: {original}")


def _validate_image_upload(upload: UploadFile) -> str:
    content_type = (upload.content_type or "").lower()
    extension = _extension_for_upload(upload)
    if content_type and not content_type.startswith("image/") and content_type != "application/octet-stream":
        raise HTTPException(status_code=400, detail=f"Il file {upload.filename} non sembra un'immagine")
    return extension


def _stored_path_to_abs(stored_path: str) -> Path:
    root = UPLOAD_ROOT.resolve()
    target = (UPLOAD_ROOT / stored_path).resolve()
    try:
        target.relative_to(root)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Percorso allegato non valido") from exc
    return target


def _attachment_payload(row: dict) -> dict:
    return {
        "id": row["id"],
        "original_filename": row.get("original_filename", ""),
        "content_type": row.get("content_type", ""),
        "size_bytes": row.get("size_bytes", 0),
        "uploaded_at": row.get("uploaded_at", ""),
        "url": f"/admin/api/attachments/{row['id']}/file",
    }


def _parse_conversation_history(value) -> list[dict]:
    if isinstance(value, list):
        return value
    if not value:
        return []
    try:
        parsed = json.loads(value)
    except (TypeError, ValueError, json.JSONDecodeError):
        return []
    return parsed if isinstance(parsed, list) else []


def _claim_session_attachments(session_id: str, complaint_id: int) -> int:
    pending = get_pending_attachments(session_id)
    if not pending:
        return 0

    claimed = 0
    dest_dir = UPLOAD_ROOT / "complaints" / str(complaint_id)
    dest_dir.mkdir(parents=True, exist_ok=True)

    for attachment in pending:
        old_abs = _stored_path_to_abs(attachment["stored_path"])
        filename = Path(attachment["stored_path"]).name
        new_abs = dest_dir / filename
        new_rel = Path("complaints") / str(complaint_id) / filename
        if old_abs.exists() and old_abs != new_abs:
            shutil.move(str(old_abs), str(new_abs))
        attach_pending_attachment(attachment["id"], complaint_id, new_rel.as_posix())
        claimed += 1

    return claimed


def _delete_complaint_upload_files(complaint_id: int, attachments: list[dict]) -> int:
    deleted = 0
    for attachment in attachments:
        stored_path = attachment.get("stored_path")
        if not stored_path:
            continue
        path = _stored_path_to_abs(stored_path)
        if path.exists() and path.is_file():
            path.unlink()
            deleted += 1

    root = UPLOAD_ROOT.resolve()
    complaint_dir = (UPLOAD_ROOT / "complaints" / str(complaint_id)).resolve()
    try:
        complaint_dir.relative_to(root)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Percorso upload reclamo non valido") from exc

    if complaint_dir.exists():
        shutil.rmtree(complaint_dir)
    return deleted


# ── Pydantic models ───────────────────────────────────────────────────────────
class ChatRequest(BaseModel):
    session_id: str
    message: str
    history: list[dict] = []

class ResetRequest(BaseModel):
    session_id: str

class PrefillRequest(BaseModel):
    session_id: str
    name: str = ""
    email: str = ""

class ComplaintUpdate(BaseModel):
    status: str | None = None
    ai_response: str | None = None
    closed_at: str | None = None
    priority: str | None = None

class ConfigPayload(BaseModel):
    common_knowledge: str = ""
    classification_rules: str = ""
    anthropic_api_key: str = ""
    model: str = "claude-haiku-4-5-20251001"
    clusters: list[dict] | None = None
    products: list[str] | None = None


# ══════════════════════════════════════════════════════════════════════════════
# PUBLIC CHAT API
# ══════════════════════════════════════════════════════════════════════════════
@app.post("/api/chat")
async def chat(req: ChatRequest):
    try:
        state = get_session(req.session_id)
        reply, new_state, suggestions = process_message(req.message, state, req.history)
        attachments_count = 0
        if new_state.get("phase") == "done" and new_state.get("complaint_id"):
            attachments_count = _claim_session_attachments(req.session_id, new_state["complaint_id"])
            if attachments_count:
                new_state.setdefault("collected", {})["has_photo"] = True
            try:
                update_complaint(new_state["complaint_id"], {
                    "conversation_history": json.dumps(
                        build_conversation_history(req.history, req.message, reply),
                        ensure_ascii=False,
                    )
                })
            except Exception:
                logger.warning(
                    "Unable to persist conversation history for complaint %s",
                    new_state["complaint_id"],
                    exc_info=True,
                )
        set_session(req.session_id, new_state)
        collected = new_state.get("collected", {})
        return {
            "reply": reply,
            "suggestions": suggestions,
            "phase": new_state.get("phase"),
            "show_upload": new_state.get("phase") in ("collecting", "done"),
            "complaint_id": new_state.get("complaint_id"),
            "attachments_count": attachments_count,
            "customer_name":  collected.get("name")  if new_state.get("phase") == "done" else None,
            "customer_email": collected.get("email") if new_state.get("phase") == "done" else None,
        }
    except Exception as e:
        logger.error("Chat error: %s", e, exc_info=True)
        return JSONResponse(status_code=500, content={"error": "Errore interno. Riprova."})


@app.post("/api/chat/uploads")
async def upload_chat_images(
    session_id: str = Form(...),
    complaint_id: int | None = Form(None),
    files: list[UploadFile] = File(...),
):
    if not files:
        raise HTTPException(status_code=400, detail="Nessuna immagine ricevuta")
    if len(files) > MAX_UPLOAD_FILES:
        raise HTTPException(status_code=400, detail=f"Puoi caricare al massimo {MAX_UPLOAD_FILES} immagini alla volta")

    state = get_session(session_id)
    target_complaint_id = complaint_id or state.get("complaint_id")
    if target_complaint_id and not get_complaint_by_id(target_complaint_id):
        raise HTTPException(status_code=404, detail="Reclamo non trovato")

    prepared = []
    for upload in files:
        extension = _validate_image_upload(upload)
        content = await upload.read(MAX_IMAGE_BYTES + 1)
        if len(content) > MAX_IMAGE_BYTES:
            raise HTTPException(status_code=413, detail=f"{upload.filename} supera il limite di 8 MB")
        if not content:
            raise HTTPException(status_code=400, detail=f"{upload.filename} e' vuoto")
        prepared.append((upload, extension, content))

    if target_complaint_id:
        base_rel = Path("complaints") / str(target_complaint_id)
    else:
        base_rel = Path("pending") / _safe_session_dir(session_id)
    base_abs = UPLOAD_ROOT / base_rel
    base_abs.mkdir(parents=True, exist_ok=True)

    attachments = []
    for upload, extension, content in prepared:
        stored_name = f"{uuid.uuid4().hex}{extension}"
        rel_path = base_rel / stored_name
        abs_path = UPLOAD_ROOT / rel_path
        abs_path.write_bytes(content)

        attachment_id = save_attachment_record(
            session_id=session_id,
            complaint_id=target_complaint_id,
            original_filename=_safe_original_filename(upload.filename),
            stored_path=rel_path.as_posix(),
            content_type=upload.content_type or mimetypes.guess_type(str(abs_path))[0] or "image/*",
            size_bytes=len(content),
        )
        attachments.append(_attachment_payload({
            "id": attachment_id,
            "original_filename": _safe_original_filename(upload.filename),
            "content_type": upload.content_type or "",
            "size_bytes": len(content),
            "uploaded_at": "",
        }))

    if target_complaint_id:
        state["complaint_id"] = target_complaint_id
        state.setdefault("collected", {})["has_photo"] = True
        set_session(session_id, state)
    else:
        state.setdefault("collected", {})["has_photo"] = True
        if state.get("waiting_for") == "photo":
            state["waiting_for"] = None
        set_session(session_id, state)

    return {"ok": True, "count": len(attachments), "attachments": attachments}


@app.post("/api/session/reset")
async def reset_session(req: ResetRequest):
    set_session(req.session_id, init_state())
    return {"ok": True}


@app.post("/api/session/prefill")
async def prefill_session(req: PrefillRequest):
    from modules.constants import REQUIRED_COMPLAINT_FIELDS
    state = get_session(req.session_id)
    collected = state.get("collected", {})
    if req.name:  collected["name"]  = req.name
    if req.email: collected["email"] = req.email
    state["collected"] = collected
    state["phase"] = "collecting"
    missing = [f for f in REQUIRED_COMPLAINT_FIELDS if not collected.get(f)]
    state["waiting_for"] = missing[0] if missing else None
    set_session(req.session_id, state)
    return {"ok": True}


@app.get("/api/status")
async def status():
    config = get_chatbot_config()
    return {
        "llm_configured": llm_module.is_configured(config),
        "active_sessions": len(_sessions),
    }


# ══════════════════════════════════════════════════════════════════════════════
# ADMIN API
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/admin/api/complaints")
async def admin_list_complaints(
    status: str = Query("", alias="status"),
    category: str = Query("", alias="category"),
    search: str = Query("", alias="search"),
    limit: int = 100,
    offset: int = 0,
):
    filters = {}
    if status:   filters["status"]   = status
    if category: filters["category"] = category
    if search:   filters["search"]   = search
    rows = get_complaints(filters)
    total = len(rows)
    return {"total": total, "items": rows[offset: offset + limit]}


@app.get("/admin/api/complaints/{complaint_id}")
async def admin_get_complaint(complaint_id: int):
    row = get_complaint_by_id(complaint_id)
    if not row:
        return JSONResponse(status_code=404, content={"error": "Non trovato"})
    row["attachments"] = [
        _attachment_payload(attachment)
        for attachment in get_complaint_attachments(complaint_id)
    ]
    row["conversation_history"] = _parse_conversation_history(row.get("conversation_history"))
    return row


@app.get("/admin/api/attachments/{attachment_id}/file")
async def admin_get_attachment_file(attachment_id: int):
    attachment = get_attachment_by_id(attachment_id)
    if not attachment or not attachment.get("complaint_id"):
        return JSONResponse(status_code=404, content={"error": "Allegato non trovato"})

    path = _stored_path_to_abs(attachment["stored_path"])
    if not path.exists():
        return JSONResponse(status_code=404, content={"error": "File allegato non trovato"})

    return FileResponse(
        path,
        media_type=attachment.get("content_type") or "application/octet-stream",
    )


@app.patch("/admin/api/complaints/{complaint_id}")
async def admin_update_complaint(complaint_id: int, payload: ComplaintUpdate):
    data = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not data:
        return {"ok": True}
    # auto-set closed_at when closing
    if data.get("status") in ("Chiuso", "Chiuso automaticamente") and "closed_at" not in data:
        data["closed_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    update_complaint(complaint_id, data)
    return {"ok": True}


@app.delete("/admin/api/complaints/{complaint_id}")
async def admin_delete_complaint(complaint_id: int):
    row = get_complaint_by_id(complaint_id)
    if not row:
        return JSONResponse(status_code=404, content={"error": "Reclamo non trovato"})

    attachments = get_complaint_attachments(complaint_id)
    try:
        deleted_files = _delete_complaint_upload_files(complaint_id, attachments)
    except OSError as exc:
        logger.error("Unable to delete upload files for complaint %s: %s", complaint_id, exc, exc_info=True)
        raise HTTPException(status_code=500, detail="Impossibile eliminare tutti gli allegati del reclamo") from exc

    deleted = delete_complaint(complaint_id)
    if not deleted:
        return JSONResponse(status_code=404, content={"error": "Reclamo non trovato"})
    return {"ok": True, "deleted_files": deleted_files}


@app.get("/admin/api/stats")
async def admin_stats(
    date_from: str = Query("", alias="date_from"),
    date_to:   str = Query("", alias="date_to"),
    product:   str = Query("", alias="product"),
    category:  str = Query("", alias="category"),
):
    rows = get_stats()

    # Parse and filter
    def parse_dt(s):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try: return datetime.strptime(s, fmt)
            except: pass
        return None

    if date_from:
        df = parse_dt(date_from)
        if df: rows = [r for r in rows if parse_dt(r.get("created_at","") or "") and parse_dt(r["created_at"]) >= df]
    if date_to:
        dt = parse_dt(date_to + " 23:59:59")
        if dt: rows = [r for r in rows if parse_dt(r.get("created_at","") or "") and parse_dt(r["created_at"]) <= dt]
    if product:  rows = [r for r in rows if r.get("product")           == product]
    if category: rows = [r for r in rows if r.get("problem_category")  == category]

    total   = len(rows)
    open_c  = sum(1 for r in rows if r.get("status") == "Aperto")
    auto_c  = sum(1 for r in rows if r.get("status") == "Chiuso automaticamente")
    pending = sum(1 for r in rows if r.get("status") in ("Aperto", "In lavorazione"))

    # Avg closure days
    closed = [r for r in rows if r.get("closed_at")]
    avg_days = None
    if closed:
        deltas = []
        for r in closed:
            c = parse_dt(r["created_at"])
            cl = parse_dt(r["closed_at"])
            if c and cl: deltas.append((cl - c).days)
        avg_days = round(sum(deltas) / len(deltas), 1) if deltas else None

    top_cat = None
    top_product = None
    top_cluster1 = None
    if rows:
        from collections import Counter
        counts = Counter(r.get("problem_category","") for r in rows if r.get("problem_category"))
        top_cat = counts.most_common(1)[0][0] if counts else None
        product_counts = Counter(r.get("product","") for r in rows if r.get("product"))
        top_product = product_counts.most_common(1)[0][0] if product_counts else None
        cluster1_counts = Counter(r.get("cluster1","") for r in rows if r.get("cluster1"))
        top_cluster1 = cluster1_counts.most_common(1)[0][0] if cluster1_counts else None

    # Monthly trend (last 24 months)
    from collections import defaultdict
    monthly = defaultdict(int)
    for r in rows:
        dt = parse_dt(r.get("created_at","") or "")
        if dt:
            key = dt.strftime("%Y-%m")
            monthly[key] += 1
    # Fill missing months
    now = datetime.now()
    all_months = [(now - timedelta(days=30*i)).strftime("%Y-%m") for i in range(23,-1,-1)]
    monthly_series = [{"month": m, "count": monthly.get(m, 0)} for m in all_months]

    # By category
    cat_counts = defaultdict(int)
    for r in rows:
        cat_counts[r.get("problem_category","Altro")] += 1
    by_category = [{"label": k, "value": v} for k, v in sorted(cat_counts.items(), key=lambda x: -x[1])]

    # By product
    prod_counts = defaultdict(int)
    for r in rows:
        prod_counts[r.get("product","—")] += 1
    by_product = [{"label": k, "value": v} for k, v in sorted(prod_counts.items(), key=lambda x: -x[1])[:10]]

    # Auto vs manual
    auto_count   = sum(1 for r in rows if r.get("classification") == "semplice")
    manual_count = sum(1 for r in rows if r.get("classification") == "complesso")
    auto_vs_manual = [
        {"label": "Risposta automatica", "value": auto_count},
        {"label": "Gestito dal team",    "value": manual_count},
    ]

    # Avg days to close by category
    cat_days = defaultdict(list)
    for r in closed:
        cat = r.get("problem_category","Altro")
        c  = parse_dt(r["created_at"])
        cl = parse_dt(r["closed_at"])
        if c and cl: cat_days[cat].append((cl - c).days)
    avg_by_cat = [
        {"label": k, "value": round(sum(v)/len(v), 1)}
        for k, v in sorted(cat_days.items(), key=lambda x: -sum(x[1])/len(x[1]))
        if v
    ]

    # Top 5 products
    top5 = sorted(prod_counts.items(), key=lambda x: -x[1])[:5]

    return {
        "kpis": {
            "total":    total,
            "open":     open_c,
            "auto":     auto_c,
            "pending":  pending,
            "avg_days": avg_days,
            "top_cat":  top_cat,
            "top_product": top_product,
            "top_cluster1": top_cluster1,
        },
        "monthly":        monthly_series,
        "by_category":    by_category,
        "by_product":     by_product,
        "auto_vs_manual": auto_vs_manual,
        "avg_by_cat":     avg_by_cat,
        "top5_products":  [{"product": k, "count": v} for k, v in top5],
        "unique_products": sorted(set(r.get("product","") for r in get_stats() if r.get("product"))),
        "unique_categories": sorted(set(r.get("problem_category","") for r in get_stats() if r.get("problem_category"))),
    }


@app.get("/admin/api/export")
async def admin_export():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("""
        SELECT id, created_at, customer_name, customer_email, product,
               cluster1, cluster2, gravity, problem_category,
               lot_code, expiry_date, purchase_location, description,
               status, classification, auto_response, ai_response,
               closed_at, has_photo
        FROM complaints ORDER BY id DESC
    """).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reclami"

    headers = [
        "ID", "Data apertura", "Nome cliente", "Email", "Prodotto",
        "Cluster 1", "Cluster 2", "Gravità", "Categoria (legacy)",
        "Codice lotto", "Data scadenza", "Punto vendita",
        "Descrizione", "Stato", "Classificazione", "Risposta automatica",
        "Risposta inviata", "Data chiusura", "Foto allegata",
    ]

    header_fill = PatternFill("solid", fgColor="C0272D")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    col_widths = [6, 18, 22, 28, 20, 20, 30, 10, 22, 12, 14, 22, 50, 18, 14, 18, 60, 18, 14]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    for r, row in enumerate(rows, 2):
        ws.cell(r, 1,  row["id"])
        ws.cell(r, 2,  row["created_at"])
        ws.cell(r, 3,  row["customer_name"])
        ws.cell(r, 4,  row["customer_email"])
        ws.cell(r, 5,  row["product"])
        ws.cell(r, 6,  row["cluster1"])
        ws.cell(r, 7,  row["cluster2"])
        ws.cell(r, 8,  row["gravity"] or row["priority"])
        ws.cell(r, 9,  row["problem_category"])
        ws.cell(r, 10, row["lot_code"])
        ws.cell(r, 11, row["expiry_date"])
        ws.cell(r, 12, row["purchase_location"])
        ws.cell(r, 13, row["description"])
        ws.cell(r, 14, row["status"])
        ws.cell(r, 15, row["classification"])
        ws.cell(r, 16, "Sì" if row["auto_response"] else "No")
        ws.cell(r, 17, row["ai_response"])
        ws.cell(r, 18, row["closed_at"])
        ws.cell(r, 19, "Sì" if row["has_photo"] else "No")
        for col in (13, 17):
            ws.cell(r, col).alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"reclami_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/admin/api/config")
async def admin_get_config():
    return get_chatbot_config()


@app.post("/admin/api/config")
async def admin_save_config(payload: ConfigPayload):
    import os
    existing = get_chatbot_config()
    incoming = payload.model_dump()
    cfg = {
        **existing,
        **{key: value for key, value in incoming.items() if key in payload.model_fields_set},
    }
    save_chatbot_config(cfg)
    if cfg.get("anthropic_api_key"):
        os.environ["ANTHROPIC_API_KEY"] = cfg["anthropic_api_key"]
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════════
# STATIC & HTML
# ══════════════════════════════════════════════════════════════════════════════
app.mount("/assets", StaticFiles(directory="assets"), name="assets")

@app.get("/", response_class=HTMLResponse)
async def index():
    return FileResponse("templates/index.html")

@app.get("/admin", response_class=HTMLResponse)
async def admin():
    return FileResponse("templates/admin.html")
